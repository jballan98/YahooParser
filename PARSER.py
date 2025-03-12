import os
import yfinance as yf
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox
import pandas as pd  # Import pandas

# Create a Tkinter root window once to reuse for message boxes
root = tk.Tk()
root.withdraw()  # Hide the root window

# Function to show success message
def show_success_message():
    messagebox.showinfo("Success", "The process has completed successfully!")  # Show the success message
    root.quit()  # Quit the Tkinter event loop

# Function to show error message
def show_error_message(error_message):
    messagebox.showerror("Error", f"An error occurred: {error_message}")  # Show the error message
    root.quit()  # Quit the Tkinter event loop

# Create 'info' folder if it doesn't exist
folder_path = './info'
os.makedirs(folder_path, exist_ok=True)

# Fetch the ticker symbol
ticker = 'AAPL'  # You can dynamically change this as needed
# ticker = input("Enter Ticker: ").strip().upper()
stock_data = yf.Ticker(ticker)

try:
    # Fetch the historical data (price data) for the ticker
    historical_data = stock_data.history(period="5y")

    # Check if data is returned
    if historical_data.empty:
        raise ValueError(f"No price data found for {ticker}. The stock might be delisted or unavailable.")

    # Check if the datetime index is timezone-aware and remove timezone if present
    if historical_data.index.tz is not None:
        historical_data.index = historical_data.index.tz_localize(None)

    # Get financial data
    income_statement = stock_data.financials
    balance_sheet = stock_data.balance_sheet
    cash_flow = stock_data.cashflow

    # Get the info data (contains key statistics and fundamentals)
    info_data = stock_data.info  # This will pull data from the 'info' page

    # Clean the info data: convert problematic types to strings or numbers
    def clean_info_data(info_dict):
        cleaned_info = {}
        for key, value in info_dict.items():
            # If value is None, replace it with an empty string
            if value is None:
                cleaned_info[key] = ""
            # If value is a list or dict, convert to string (optional)
            elif isinstance(value, (list, dict)):
                cleaned_info[key] = str(value)  # You may modify this if a deeper structure is needed
            # If value is boolean, convert to string
            elif isinstance(value, bool):
                cleaned_info[key] = str(value)
            else:
                cleaned_info[key] = value
        return cleaned_info

    # Clean the info data before converting to DataFrame
    cleaned_info_data = clean_info_data(info_data)

    # Convert the cleaned info data into a DataFrame
    info_data_df = pd.DataFrame(list(cleaned_info_data.items()), columns=['Statistic', 'Value'])

    # Define the input and output file paths
    input_file = 'DCF Model.xlsx'  # The existing Excel file in the project folder
    output_file = os.path.join(folder_path, f'{ticker}_DCF_Model.xlsx')  # Save it in the 'info' folder with the ticker

    # Try to load the workbook if it exists
    try:
        # Open the existing workbook
        book = load_workbook(input_file)
        print(f"Opened existing file: {input_file}")
    except FileNotFoundError:
        show_error_message(f"The file {input_file} does not exist. Please make sure it is in the project folder.")
        exit(1)  # Exit the script if the file is not found

    # Function to add DataFrame to an Excel sheet
    def add_dataframe_to_sheet(sheet_name, data, wb, transpose=False, delete_second_row=False, format_date=False):
        if sheet_name not in wb.sheetnames:
            sheet = wb.create_sheet(title=sheet_name)
        else:
            sheet = wb[sheet_name]  # If sheet exists, access it
            # Clear existing data (if any) by deleting all rows
            sheet.delete_rows(1, sheet.max_row)  # Deletes all rows in the sheet

        # Transpose the data if required
        if transpose:
            data = data.transpose()

        # Add new data to the sheet
        for row in dataframe_to_rows(data, index=True, header=True):
            sheet.append(row)

        # After adding data, delete the second row if required
        if delete_second_row:
            sheet.delete_rows(2)

        # Format the top row to short date format
        if format_date:
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col)
                if isinstance(cell.value, str) and cell.value in data.columns:
                    try:
                        # Try to parse date if column header looks like a date
                        cell.number_format = 'mm/dd/yyyy'  # Excel short date format
                    except ValueError:
                        pass  # If the value is not a date, we just skip it

    # Add historical data (price data) without transposing
    add_dataframe_to_sheet('Historical Data', historical_data, book, transpose=False)

    # Add Income Statement (transpose, delete second row)
    add_dataframe_to_sheet('Income Statement', income_statement.transpose(), book, transpose=True, delete_second_row=True, format_date=True)

    # Add Balance Sheet (transpose, delete second row)
    add_dataframe_to_sheet('Balance Sheet', balance_sheet.transpose(), book, transpose=True, delete_second_row=True, format_date=True)

    # Add Cash Flow (transpose, delete second row)
    add_dataframe_to_sheet('Cash Flow', cash_flow.transpose(), book, transpose=True, delete_second_row=True, format_date=True)

    # Add Info Data (statistics from the 'info' page)
    add_dataframe_to_sheet('Info', info_data_df, book, transpose=False, delete_second_row=True)

    # Save the modified workbook to the 'info' folder (with ticker in filename)
    book.save(output_file)
    print(f"Excel file saved at {output_file}")

    # Show success message
    show_success_message()

except ValueError as e:
    show_error_message(str(e))
    exit(1)  # Exit if no price data is found

except Exception as e:
    show_error_message(f"An unexpected error occurred: {str(e)}")
    exit(1)  # Exit in case of other errors
